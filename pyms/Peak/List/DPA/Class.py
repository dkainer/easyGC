"""
Classes for peak alignment by dynamic programming
"""

 #############################################################################
 #                                                                           #
 #    PyMS software for processing of metabolomic mass-spectrometry data     #
 #    Copyright (C) 2005-2012 Vladimir Likic                                 #
 #                                                                           #
 #    This program is free software; you can redistribute it and/or modify   #
 #    it under the terms of the GNU General Public License version 2 as      #
 #    published by the Free Software Foundation.                             #
 #                                                                           #
 #    This program is distributed in the hope that it will be useful,        #
 #    but WITHOUT ANY WARRANTY; without even the implied warranty of         #
 #    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the          #
 #    GNU General Public License for more details.                           #
 #                                                                           #
 #    You should have received a copy of the GNU General Public License      #
 #    along with this program; if not, write to the Free Software            #
 #    Foundation, Inc., 675 Mass Ave, Cambridge, MA 02139, USA.              #
 #                                                                           #
 #############################################################################

import copy

import numpy, Pycluster

from pyms.Utils.Error import error, stop
from pyms.Experiment.Class import Experiment
from pyms.GCMS.Class import MassSpectrum
from pyms.Peak.Class import Peak
from pyms.Peak.List.Function import composite_peak

import Function

#dk
import math
import operator
from pyms.Peak.List.Function import percentile_based_outlier, mad_based_outlier, median_outliers
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl import utils

# If psyco is installed, use it to speed up running time
try:
    import psyco
    psyco.full()
except:
    pass

class Alignment(object):

    """
    @summary: Models an alignment of peak lists

    @author: Woon Wai Keen
    @author: Vladimir Likic
    """

    def __init__(self, expr):

        """
        @param expr: The experiment to be converted into an alignment object
        @type expr: pyms.Experiment.Class.Experiment

        @author: Woon Wai Keen
        @author: Qiao Wang
        @author: Vladimir Likic
        """
        if expr == None:
            self.peakpos = []
            self.peakalgt = []
            self.expr_code =  []
            self.similarity = None
        else:
            if not isinstance(expr, Experiment):
                error("'expr' must be an Experiment object")
            #for peak in expr.get_peak_list():
            #    if peak.get_area() == None or peak.get_area() <= 0:
            #        error("All peaks must have an area for alignment")
            self.peakpos = [ copy.deepcopy(expr.get_peak_list()) ]
            self.peakalgt = numpy.transpose(self.peakpos)
            self.expr_code =  [ expr.get_expr_code() ]
            self.similarity = None

    def __len__(self):

        """
        @summary: Returns the length of the alignment, defined as the number of
        peak positions in the alignment

        @author: Qiao Wang
        @author: Vladimir Likic
        """

        return len(self.peakalgt)

    def filter_min_peaks(self, min_peaks):

        """
        @summary: Filters alignment positions that have less peaks than 'min_peaks'

        This function is useful only for within state alignment.

        @param min_peaks: Minimum number of peaks required for the alignment
            position to survive filtering
        @type min_peaks: IntType

        @author: Qiao Wang
        """

        filtered_list=[]

        for pos in range(len(self.peakalgt)):
            if len(filter(None,self.peakalgt[pos])) >= min_peaks:
                filtered_list.append(self.peakalgt[pos])

        self.peakalgt = filtered_list
        self.peakpos = numpy.transpose(self.peakalgt)



    def write_csv_dk(self, rt_file_name, area_file_name, minutes=True):

        """
        @summary: Writes the alignment to CSV files, but excluded outliers from the calculation of composite peak

        This function writes two files: one containing the alignment of peak
        retention times and the other containing the alignment of peak areas.

        @param rt_file_name: The name for the retention time alignment file
        @type rt_file_name: StringType
        @param area_file_name: The name for the areas alignment file
        @type area_file_name: StringType
        @param minutes: An optional indicator whether to save retention times
            in minutes. If False, retention time will be saved in seconds
        @type minutes: BooleanType

        @author: David Kainer
        """

        try:
            fp1 = open(rt_file_name, "w")
            fp2 = open(area_file_name, "w")
        except IOError:
            error("Cannot open output file for writing")

        # create header
        header = '"UID","RTavg"'
        for item in self.expr_code:
            expr_code = ( '"%s"' % item )
            header = header + "," + expr_code
        header = header + "\n"

        # write headers
        fp1.write(header)
        fp2.write(header)

        # for each alignment position write alignment's peak and area
        for peak_idx in range(len(self.peakpos[0])):    # loop through peak lists (rows)

            rts = []
            areas = []
            new_peak_list = []

            for align_idx in range(len(self.peakpos)):   # loops through samples (columns)
                peak = self.peakpos[align_idx][peak_idx]

                if peak is not None:

                    if minutes:
                        rt = peak.get_rt()/60.0
                    else:
                        rt = peak.get_rt()

                    rts.append(rt)
                    areas.append(peak.get_area())
                    new_peak_list.append(peak)

                else:
                    rts.append(numpy.nan)
                    areas.append(None)

            compo_peak = composite_peak(new_peak_list, minutes)
            peak_UID = compo_peak.get_UID()
            peak_UID_string = ( '"%s"' % peak_UID)

            # write to retention times file
            fp1.write(peak_UID_string)
            fp1.write(",%.3f" % float(compo_peak.get_rt()/60))

            for rt in rts:
                if numpy.isnan(rt):
                    fp1.write(",NA")
                else:
                    fp1.write(",%.3f" % rt)
            fp1.write("\n")

            # write to peak areas file
            fp2.write(peak_UID_string)
            fp2.write(",%.3f" % float(compo_peak.get_rt()/60))
            for area in areas:
                if area == None:
                    fp2.write(",NA")
                else:
                    fp2.write(",%.0f" % area)
            fp2.write("\n")

        fp1.close()
        fp2.close()


    def write_excel(self, excel_file_name, minutes=True):
        """
        @summary: Writes the alignment to an excel file, with colouring showing possible mis-alignments

        @param excel_file_name: The name for the retention time alignment file
        @type excel_file_name: StringType
        @param minutes: An optional indicator whether to save retention times
            in minutes. If False, retention time will be saved in seconds
        @type minutes: BooleanType

        @author: David Kainer
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Aligned RT"

        # create header row
        ws['A1'] = "UID"
        ws['B1'] = "RTavg"
        for i,item in enumerate(self.expr_code):
            currcell = ws.cell( row = 1, column = i+3, value= "%s" % item )
            comment = Comment('sample '+str(i), 'dave')
            currcell.comment = comment

        # for each alignment position write alignment's peak and area
        for peak_idx in range(len(self.peakpos[0])):    # loop through peak lists (rows)

            new_peak_list = []

            for align_idx in range(len(self.peakpos)):   # loops through samples (columns)
                peak = self.peakpos[align_idx][peak_idx]

                if peak is not None:

                    if minutes:
                        rt = peak.get_rt()/60.0
                    else:
                        rt = peak.get_rt()

                    area = peak.get_area()
                    new_peak_list.append(peak)

                    # write the RT into the cell in the excel file
                    currcell = ws.cell( row = 2+peak_idx, column = 3+align_idx, value=round(rt, 3) )

                    # get the mini-mass spec for this peak, and divide the ion intensities by 1000 to shorten them
                    ia = peak.get_ion_areas()
                    ia.update( (mass, int(intensity/1000)) for mass, intensity in ia.items() )
                    sorted_ia = sorted(ia.iteritems(), key=operator.itemgetter(1), reverse=True)

                    # write the peak area and mass spec into the comment for the cell
                    comment = Comment("Area: %.0f | MassSpec: %s" % (area,sorted_ia), 'dave')
                    currcell.number_format
                    currcell.comment = comment

                else:
                    rt = 'NA'
                    area = 'NA'
                    currcell = ws.cell( row = 2+peak_idx, column = 3+align_idx, value='NA' )
                    comment = Comment("Area: NA", 'dave')
                    currcell.number_format
                    currcell.comment = comment

            compo_peak      = composite_peak(new_peak_list, minutes)
            peak_UID        = compo_peak.get_UID()
            peak_UID_string = ( '"%s"' % peak_UID)

            currcell = ws.cell( row = 2+peak_idx, column = 1, value = peak_UID_string )
            currcell = ws.cell( row = 2+peak_idx, column = 2, value = "%.3f" % float(compo_peak.get_rt()/60) )

        # colour the cells in each row based on their RT percentile for that row
        i = 0
        for row in ws.rows:
            i += 1
            cell_range = ("{0}"+str(i)+":{1}"+str(i)).format(utils.get_column_letter(3), utils.get_column_letter(len(row)))
            ws.conditional_formatting.add(cell_range, ColorScaleRule(start_type='percentile', start_value=1, start_color='E5FFCC',
                                                               mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                                                               end_type='percentile', end_value=99, end_color='FFE5CC'))
        wb.save(excel_file_name)







    def write_csv(self, rt_file_name, area_file_name, minutes=True):

        """
        @summary: Writes the alignment to CSV files

        This function writes two files: one containing the alignment of peak
        retention times and the other containing the alignment of peak areas.

        @param rt_file_name: The name for the retention time alignment file
        @type rt_file_name: StringType
        @param area_file_name: The name for the areas alignment file
        @type area_file_name: StringType
        @param minutes: An optional indicator whether to save retention times
            in minutes. If False, retention time will be saved in seconds
        @type minutes: BooleanType

        @author: Woon Wai Keen
        @author: Andrew Isaac
        @author: Vladimir Likic
        """

        try:
            fp1 = open(rt_file_name, "w")
            fp2 = open(area_file_name, "w")
        except IOError:
            error("Cannot open output file for writing")

        # create header
        header = '"UID","RTavg"'
        for item in self.expr_code:
            expr_code = ( '"%s"' % item )
            header = header + "," + expr_code
        header = header + "\n"

        # write headers
        fp1.write(header)
        fp2.write(header)

        # for each alignment position write alignment's peak and area
        for peak_idx in range(len(self.peakpos[0])):

            rts = []
            areas = []
            new_peak_list = []
            avgrt = 0
            countrt = 0

            for align_idx in range(len(self.peakpos)):

                peak = self.peakpos[align_idx][peak_idx]

                if peak is not None:

                    if minutes:
                        rt = peak.get_rt()/60.0
                    else:
                        rt = peak.get_rt()

                    rts.append(rt)
                    areas.append(peak.get_area())
                    new_peak_list.append(peak)

                    avgrt = avgrt + rt
                    countrt = countrt + 1
                else:
                    rts.append(None)
                    areas.append(None)

            if countrt > 0:
                avgrt = avgrt/countrt

            compo_peak = composite_peak(new_peak_list, minutes)
            peak_UID = compo_peak.get_UID()
            peak_UID_string = ( '"%s"' % peak_UID)

            # write to retention times file
            fp1.write(peak_UID_string)
            fp1.write(",%.3f" % avgrt)
            
            for rt in rts:
                if rt == None:
                    fp1.write(",NA")
                else:
                    fp1.write(",%.3f" % rt)
            fp1.write("\n")

            # write to peak areas file
            fp2.write(peak_UID_string)
            fp2.write(",%.3f" % avgrt)
            for area in areas:
                if area == None:
                    fp2.write(",NA")
                else:
                    fp2.write(",%.4f" % area)
            fp2.write("\n")

        fp1.close()
        fp2.close()

    def write_common_ion_csv(self, area_file_name, top_ion_list, minutes=True):

        """
        @summary: Writes the alignment to CSV files

        This function writes two files: one containing the alignment of peak
        retention times and the other containing the alignment of peak areas.

        @param area_file_name: The name for the areas alignment file
        @type area_file_name: StringType
        @param top_ion_list: A list of the highest intensity common ion
                             along the aligned peaks
        @type top_ion_list: ListType
        @param minutes: An optional indicator whether to save retention times
            in minutes. If False, retention time will be saved in seconds
        @type minutes: BooleanType

        @author: Woon Wai Keen
        @author: Andrew Isaac
        @author: Sean O'Callaghan
        @author: Vladimir Likic
        """

        try:
            fp = open(area_file_name, "w")
        except IOError:
            error("Cannot open output file for writing")

        if top_ion_list == None:
            error("List of common ions must be supplied")

        # create header
        header = '"UID","RTavg", "Quant Ion"'
        for item in self.expr_code:
            expr_code = ( '"%s"' % item )
            header = header + "," + expr_code
        header = header + "\n"

        # write headers

        fp.write(header)

        rtsums = []
        rtcounts = []

        # The following two arrays will become list of lists
        # such that:
        # areas = [  [align1_peak1, align2_peak1, .....,alignn_peak1]
        #            [align1_peak2, ................................]
        #              .............................................
        #            [align1_peakm,....................,alignn_peakm]  ]
        areas = []
        new_peak_lists = []

        for peak_list in self.peakpos:
            index = 0
            for peak in peak_list:
                # one the first iteration, populate the lists
                if len(areas) < len(peak_list):
                    areas.append([])
                    new_peak_lists.append([])
                    rtsums.append(0)
                    rtcounts.append(0)

                if peak is not None:
                    rt = peak.get_rt()

                    # get the area of the common ion for the peak
                    # an area of 'na' shows that while the peak was
                    # aligned, the common ion was not present
                    area = peak.get_ion_area(top_ion_list[index])
                     
                    areas[index].append(area)
                    new_peak_lists[index].append(peak)

                    # The following code to the else statement is
                    # just for calculating the average rt
                    rtsums[index] += rt
                    rtcounts[index] += 1
                    
                else:
                    areas[index].append(None)

                index += 1

        out_strings = []
        index = 0
        # now write the strings for the file
        for area_list in areas:
 
            # write initial info:
            # peak unique id, peak average rt
            compo_peak = composite_peak(new_peak_lists[index], minutes)
            peak_UID = compo_peak.get_UID()
            peak_UID_string = ( '"%s"' % peak_UID)

            rt_avg = rtsums[index]/rtcounts[index]
                    
            out_strings.append(peak_UID_string + (",%.3f" % (rt_avg/60))+\
                                   (",%d" % top_ion_list[index]))

            for area in area_list:
                if area is not None:
                    out_strings[index] += (",%.4f" % area)
                else:
                    out_strings[index] += (",NA")
            
            index += 1

        # now write the file
#        print "length of areas[0]", len(areas[0])
#        print "lenght of areas", len(areas)
#        print "length of out_strings", len(out_strings)
        for row in out_strings:
            fp.write(row +"\n")
                
        fp.close()

    # written by dkainer
    def write_ion_areas_csv(self, ms_file_name, minutes=True):
        try:
            fp1 = open(ms_file_name, "w")   #dk
        except IOError:
            error("Cannot open output file for writing")
        
        # create header
        header = '"UID"|"RTavg"'
        for item in self.expr_code:
            expr_code = ( '"%s"' % item )
            header = header + "|" + expr_code
        header = header + "\n"
        
        fp1.write(header)                   #dk
        
        for peak_idx in range(len(self.peakpos[0])):

            rts = []
            ias = []
            new_peak_list = []
            avgrt = 0
            countrt = 0

            for align_idx in range(len(self.peakpos)):

                peak = self.peakpos[align_idx][peak_idx]

                if peak is not None:

                    if minutes:
                        rt = peak.get_rt()/60.0
                    else:
                        rt = peak.get_rt()

                    rts.append(rt)
                    
                    ia = peak.get_ion_areas()
                    ia.update( (mass, math.floor(intensity)) for mass, intensity in ia.items() )
                    sorted_ia = sorted(ia.iteritems(), key=operator.itemgetter(1), reverse=True)
                    ias.append(sorted_ia)
                    new_peak_list.append(peak)
                    
                    avgrt = avgrt + rt
                    countrt = countrt + 1
                else:
                    rts.append(None)
                    ias.append(None)

            if countrt > 0:
                avgrt = avgrt/countrt

            compo_peak = composite_peak(new_peak_list, minutes)
            peak_UID = compo_peak.get_UID()
            peak_UID_string = ( '"%s"' % peak_UID)

            # write to ms file
            fp1.write(peak_UID_string)
            fp1.write("|%.3f" % avgrt)
            for ia in ias:
                if ia == None:
                    fp1.write("|NA")
                else:
                    fp1.write("|%s" % ia)
            fp1.write("\n")

        fp1.close()
        
    def common_ion(self):
        """
        @summary: Calculates a common ion among the
                  peaks of an aligned peak

        @return: A list of the highest intensity common ion for all aligned 
                 peaks
        @rtype: ListType

        @author: Sean O'Callaghan

        """

        # print "#peak lists =", len(self.peakpos)
        # print "#peaks =", len(self.peakpos[0])
        
        # a list to contain the dictionaries
        # each dictionary contains the
        # top ions and their frequency for each peak
        # in the alignment
        list_of_top_ion_dicts = []
        empty_count_list = []
        
        for peak_list in self.peakpos:
            # (re)initialise the peak index
            index = 0
            
            for peak in peak_list:
                # if the dict has not been created, create it
                # will only run on first peak list
                if len(list_of_top_ion_dicts) < len(peak_list):
                    list_of_top_ion_dicts.append({})
                    
                # copy the list entry to a local dict for code clarity
                top_ion_dict = list_of_top_ion_dicts[index]

                # count the empty peaks
                if (len(empty_count_list)) < len(peak_list):
                    empty_count = 0
                else:
                    empty_count = empty_count_list[index] 
                #make sure the peak is present
                if peak is not None:
                    # get the top 5 ions
                    top_5 = peak.get_ion_areas().keys()
                
                    for ion in top_5:
                        # if we haven't seen it before, add it
                        if not top_ion_dict.has_key(ion):
                            top_ion_dict[ion]=1
                        # if we have seen it, increment the count
                        elif top_ion_dict.has_key(ion):
                            top_ion_dict[ion]+=1
                        # shouldn't happen
                        else:
                            print "error: in function common_ion()"
                else:
                    empty_count += 1
                    
                    # copy the dict back to the list
                    list_of_top_ion_dicts[index] = top_ion_dict

                    if len(empty_count_list) < len(peak_list):
                        empty_count_list.append(empty_count)
                    else:
                        empty_count_list[index] = empty_count
                # increment for the next peak
                index += 1

        #print "length of list of dicts", len(list_of_top_ion_dicts)

        #index = 0
        #for entry in list_of_top_ion_dicts:
        #    for ion in sorted(entry, key=entry.get, reverse=True)[0:3]:
        #        print ion,":", entry[ion],
        #    print '  empty:', empty_count_list[index]
        #    index += 1

        top_ion_list = []

        for entry in list_of_top_ion_dicts:
            top_ion_list.append(max(entry, key=entry.get))
            #top_ion_list.append(self.get_highest_mz_ion(entry))

        return top_ion_list
    
    def get_highest_mz_ion(self, ion_dict):
         """
         @summary: Returns the preferred ion for quantitiation
         Looks at the list of candidate ions, selects those
         which have highest occurance, and selects the heaviest
         of those
         
         @param ion_dict: a dictionary of mz value: number of occurances
         @type ion_dict: dictType
         
         @return ion: The ion to use
         @rtype: intType
         """
         max_occurances = max(ion_dict.values())
         most_freq_mzs = []
         for key, value in ion_dict.iteritems():
             if value == max_occurances:
                 most_freq_mzs.append(key)
        
         return max(most_freq_mzs)

    def aligned_peaks(self, minutes=False):

        """
        @summary: Returns a list of Peak objects where each peak
            has the combined spectra and average retention time
            of all peaks that aligned.

        @param minutes: An optional indicator of whether retention
            times are in minutes. If False, retention time are in
            seconds
        @type minutes: BooleanType

        @return: A list of composite peaks based on the alignment.
        @rtype: ListType

        @author: Andrew Isaac
        """

        # for all peaks found
        peak_list = []
        for peak_idx in range(len(self.peakpos[0])):
            # get aligned peaks, ignore missing
            new_peak_list = []
            for align_idx in range(len(self.peakpos)):
                peak = self.peakpos[align_idx][peak_idx]
                if peak is not None:
                    new_peak_list.append(peak)
            #create composite
            new_peak = composite_peak(new_peak_list, minutes)
            peak_list.append(new_peak)

        return peak_list

class PairwiseAlignment(object):

    """
    @summary: Models pairwise alignment of alignments

    @author: Woon Wai Keen
    @author: Vladimir Likic
    """

    def __init__(self, algts, D, gap):

        """
        @param algts: A list of alignments
        @type algts: ListType
        @param D: Retention time tolerance parameter for pairwise alignments
        @type D: FloatType
        @param gap: Gap parameter for pairwise alignments
        @type gap: FloatType

        @author: Woon Wai Keen
        @author: Vladimir Likic
        """
        self.algts = algts
        self.D = D
        self.gap = gap

        self.sim_matrix = self._sim_matrix(algts, D, gap)
        self.dist_matrix = self._dist_matrix(self.sim_matrix)
        self.tree = self._guide_tree(self.dist_matrix)

    def _sim_matrix(self, algts, D, gap):

        """
        @summary: Calculates the similarity matrix for the set of alignments

        @param algts: A list of alignments
        @type algts: ListType
        @param D: Retention time tolerance parameter for pairwise alignments
        @type D: FloatType
        @param gap: Gap parameter for pairwise alignments
        @type gap: FloatType

        @author: Woon Wai Keen
        @author: Vladimir Likic
        """

        n = len(algts)

        total_n = n * (n - 1) / 2

        print " Calculating pairwise alignments for %d alignments (D=%.2f, gap=%.2f)" % \
                (n, D, gap)

        sim_matrix = numpy.zeros((n,n), dtype='f')

        #DK: Could we parallelize this pairwise alignment loop??
        for i in range(n - 1):
            for j in range(i + 1, n):
                ma = Function.align(algts[i], algts[j], D, gap)
                sim_matrix[i,j] = sim_matrix[j,i] = ma.similarity
                total_n = total_n - 1
                print " -> %d pairs remaining" % total_n

        return sim_matrix

    def _dist_matrix(self, sim_matrix):

        """
        @summary: Converts similarity matrix into a distance matrix

        @param sim_matrix: The similarity matrix
        @type sim_matrix: numpy.ndarray

        @author: Woon Wai Keen
        @author: Vladimir Likic
        """

        # change similarity matrix entries (i,j) to max{matrix}-(i,j)
        sim_max = numpy.max(numpy.ravel(sim_matrix))
        dist_matrix = sim_max - sim_matrix

        # set diagonal elements of the similarity matrix to zero
        for i in range(len(dist_matrix)):
            dist_matrix[i,i] = 0

        return dist_matrix

    def _guide_tree(self, dist_matrix):

        """
        @summary: Build a guide tree from the distance matrix

        @param dist_matrix: The distance matrix
        @type dist_matrix: numpy.ndarray
        @return: Pycluster similarity tree
        @rtype: Pycluster.cluster.Tree

        @author: Woon Wai Keen
        @author: Vladimir Likic
        """

        n = len(dist_matrix)

        print " -> Clustering %d pairwise alignments." % (n*(n-1)),
        tree = Pycluster.treecluster(distancematrix=dist_matrix, method='a')
        print "Done"

        return tree

